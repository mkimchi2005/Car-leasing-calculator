from __future__ import annotations

import io
from pathlib import Path

import openpyxl
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from tax_calculator import CarOption, SalaryInput, calculate_all

# ── App setup ────────────────────────────────────────────────────────────────
app = FastAPI(title="מחשבון ליסינג")
STATIC_DIR = Path(__file__).parent / "static"
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/", response_class=HTMLResponse)
async def root():
    return (STATIC_DIR / "index.html").read_text(encoding="utf-8")


# ── Excel upload ─────────────────────────────────────────────────────────────

# Keywords for each required column — ordered from most-specific to least-specific
_COL_KEYS = {
    "manufacturer": ["שם יצרן", "שם רכב", "יצרן", "מותג", "manufacturer", "make", "brand"],
    "model":        ["שם דגם", "שם מודל", "דגם", "גרסה", "מודל", "model", "type", "סוג"],
    "cost":         ["עלות לעובד", "עלות חודשית", "תשלום עובד", "עלות עובד",
                     "עלות", "מחיר לעובד", "employee cost", "cost"],
    "benefit":      ["שווי שימוש", "שווי_שימוש", "שווי", "benefit value", "benefit", "shovi"],
}


def _detect_columns(headers: list[str]) -> dict[str, int] | None:
    """Try longest/most-specific keyword first to avoid wrong matches."""
    result: dict[str, int] = {}
    for field, keywords in _COL_KEYS.items():
        # sort keywords by length descending so "שם יצרן" beats "שם"
        for kw in sorted(keywords, key=len, reverse=True):
            for idx, h in enumerate(headers):
                if kw in h.strip():
                    if field not in result:
                        result[field] = idx
                    break
            if field in result:
                break
    required = set(_COL_KEYS)
    return result if required.issubset(result) else None


def _read_rows(ws) -> list[list]:
    """Read all data rows as plain lists (for manual mapping)."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append([str(c) if c is not None else "" for c in row])
    return rows


@app.post("/upload-cars")
async def upload_cars(file: UploadFile = File(...)):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        raw_headers = [str(c.value or "").strip()
                       for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = _detect_columns(raw_headers)

        # Always read rows so manual mapper has data if needed
        all_rows = _read_rows(ws)

        if col_map is None:
            return {
                "ok": False,
                "headers": raw_headers,
                "raw_rows": all_rows[:500],   # cap at 500 rows
                "message": (
                    f"לא זוהו אוטומטית כל העמודות הנדרשות. "
                    f"עמודות שנמצאו בקובץ: {raw_headers}. אנא בחר ידנית."
                ),
            }

        cars = []
        for row in all_rows:
            try:
                mfr  = row[col_map["manufacturer"]].strip()
                mdl  = row[col_map["model"]].strip()
                cost = float(row[col_map["cost"]]    or 0)
                ben  = float(row[col_map["benefit"]] or 0)
                if mfr or mdl:
                    cars.append({"manufacturer": mfr, "model": mdl,
                                 "employee_cost": cost, "benefit_value": ben})
            except (IndexError, ValueError):
                continue

        return {"ok": True, "cars": cars,
                "headers": raw_headers, "col_map": col_map}

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"שגיאה בקריאת הקובץ: {exc}")


# Manual column mapping endpoint
class ManualMapRequest(BaseModel):
    headers: list[str]
    rows: list[list]
    col_manufacturer: int
    col_model: int
    col_cost: int
    col_benefit: int


@app.post("/map-columns")
async def map_columns(req: ManualMapRequest):
    cars = []
    for row in req.rows:
        try:
            mfr  = str(row[req.col_manufacturer] or "").strip()
            mdl  = str(row[req.col_model]        or "").strip()
            cost = float(row[req.col_cost]    or 0)
            ben  = float(row[req.col_benefit] or 0)
            if mfr or mdl:
                cars.append({"manufacturer": mfr, "model": mdl,
                             "employee_cost": cost, "benefit_value": ben})
        except (TypeError, ValueError, IndexError):
            continue
    return {"ok": True, "cars": cars}


# ── Calculate ────────────────────────────────────────────────────────────────

class SalaryPayload(BaseModel):
    gross: float
    income_tax_actual: float = 0
    ni_actual: float = 0
    health_actual: float = 0
    social_deductions: float = 0
    net_actual: float
    credit_points: float = 2.25
    area_credit_pct: float = 0
    area_credit_cap: float = 0
    car_budget: float = 3800
    cur_car_benefit: float = 0
    cur_car_cost: float = 0
    annual_extra: float = 0


class CalculateRequest(BaseModel):
    salary: SalaryPayload
    cars: list[dict]


@app.post("/calculate")
async def calculate(req: CalculateRequest):
    sal = SalaryInput(
        gross              = req.salary.gross,
        income_tax_actual  = req.salary.income_tax_actual,
        ni_actual          = req.salary.ni_actual,
        health_actual      = req.salary.health_actual,
        social_deductions  = req.salary.social_deductions,
        net_actual         = req.salary.net_actual,
        credit_points      = req.salary.credit_points,
        area_credit_pct    = req.salary.area_credit_pct,
        area_credit_cap    = req.salary.area_credit_cap,
        car_budget         = req.salary.car_budget,
        cur_car_benefit    = req.salary.cur_car_benefit,
        cur_car_cost       = req.salary.cur_car_cost,
        annual_extra       = req.salary.annual_extra,
    )
    cars = [CarOption(**c) for c in req.cars]
    res  = calculate_all(sal, cars)

    return {
        "net_actual":      res["net_actual"],
        "net_adjustment":  res["net_adjustment"],
        "no_car_expected": res["no_car_expected"],
        "no_car_delta":    res["no_car_delta"],
        "cars": [
            {
                "manufacturer":    r.manufacturer,
                "model":           r.model,
                "employee_cost":   r.employee_cost,
                "benefit_value":   r.benefit_value,
                "expected_net":    round(r.expected_net, 2),
                "delta_vs_current":round(r.delta_vs_current, 2),
                "it_delta":        round(r.it_delta, 2),
                "ni_hi_delta":     round(r.ni_hi_delta, 2),
                "budget_effect":   round(r.budget_effect, 2),
            }
            for r in res["cars"]
        ],
    }
